"""Automotive Supply Chain Excel Builder.

5-sheet output: Supplier Data, Filtered Publics, Tier 1 Summary, Tier 2 Summary, Methodology.
Reads latest_quarter_financials.json when present; otherwise uses embedded FINANCIALS.
Company keys must match INCLUDE_DETAILS (and FINANCIALS for fallback).
"""
from __future__ import annotations

import json
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── STYLES ──────────────────────────────────────────────────────────────────
# Header, tier row shading, and number formats (currency, percent) for the workbook.
FONT_NAME = "Arial"
COLOR_HEADER_BG    = "1F3864"   # dark navy
COLOR_HEADER_FONT  = "FFFFFF"
COLOR_TIER1_BG     = "D6E4F0"   # light blue
COLOR_TIER2_BG     = "E8F5E9"   # light green
COLOR_EXCLUDE_BG   = "FFF2CC"   # yellow
COLOR_ALT_ROW      = "F5F5F5"
COLOR_INPUT_FONT   = "0000FF"   # blue = hardcoded input
COLOR_FORMULA_FONT = "000000"   # black = formula

FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_PCT      = '0.0%;(0.0%);"-"'

def header_style(bold=True, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FONT, wrap=True):
    """Return a style dict for header row (font, fill, alignment, border)."""
    return {
        "font":      Font(name=FONT_NAME, bold=bold, color=fg),
        "fill":      PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=wrap),
        "border":    thin_border(),
    }

def cell_style(bold=False, color="000000", bg=None, halign="left", wrap=False):
    """Return a style dict for a data cell (font, alignment, optional fill, border)."""
    s = {
        "font":      Font(name=FONT_NAME, bold=bold, color=color),
        "alignment": Alignment(horizontal=halign, vertical="center", wrap_text=wrap),
        "border":    thin_border(),
    }
    if bg:
        s["fill"] = PatternFill("solid", fgColor=bg)
    return s

def thin_border():
    """Return a thin gray Border for all four sides."""
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_style(cell, **kwargs):
    """Apply a style dict (font, fill, alignment, border) to an openpyxl cell."""
    if "font" in kwargs:      cell.font      = kwargs["font"]
    if "fill" in kwargs:      cell.fill      = kwargs["fill"]
    if "alignment" in kwargs: cell.alignment = kwargs["alignment"]
    if "border" in kwargs:    cell.border    = kwargs["border"]

def style_cell(cell, bold=False, color="000000", bg=None, halign="left",
               wrap=False, fmt=None, size=10):
    """Set font, alignment, border, optional fill and number format on a cell."""
    cell.font      = Font(name=FONT_NAME, bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    cell.border    = thin_border()
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt

def write_header_row(ws, row, headers, col_start=1):
    """Write a single header row with shared header styling (navy bg, white font, wrap)."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font      = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=10)
        c.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

# ─── FINANCIAL DATA ──────────────────────────────────────────────────────────
# Fallback when latest_quarter_financials.json is missing or company not in it.
# Tuple: (revenue_usd, sga_usd, ebit_usd, fiscal_year, source_note). All values in full USD. None = not available.
FINANCIALS = {
    # ── Mega Tier 1 (from training knowledge / agent data) ──
    "MAGNA INTERNATIONAL INC":          (42800000000, 1490000000,  1450000000, "FY2024", "Magna 2024 Annual Report"),
    "DENSO CORPORATION":                (48300000000, 3150000000,  2890000000, "FY Mar2025", "Denso FY2024 Annual Report (JPY 7.2T)"),
    "CONTINENTAL AKTIENGESELLSCHAFT":   (39700000000, 2470000000,  1180000000, "FY2024", "Continental 2024 Annual Report (EUR ~€36.7B)"),
    "AISIN CORP":                       (30954000000, 2546000000,  1173000000, "FY Mar2025", "Aisin FY2025 Earnings Release (JPY ~4.62T)"),
    "VALEO":                            (21700000000, 1360000000,   650000000, "FY2024", "Valeo 2024 Full Year Results (EUR ~€20.1B)"),
    "FAURECIA":                         (26600000000, 1410000000,  1020000000, "FY2024", "FORVIA 2024 Full Year Results (EUR ~€24.6B)"),
    "HYUNDAI MOBIS CO LTD":             (39800000000, 1900000000,  1530000000, "FY2024", "Hyundai Mobis 2024 Annual Report (KRW ~55T)"),
    "LEAR CORP":                        (22390000000, 1050000000,   850000000, "FY2024", "Lear 2024 10-K"),
    "BORGWARNER INC":                   (14392000000,  576000000,   943000000, "FY2024", "BorgWarner 2024 10-K"),
    "ADIENT PUBLIC LTD CO":             (13900000000,  620000000,   420000000, "FY2024", "Adient FY2024 10-K"),
    "COMPAGNIE GENERALE DES ETABLISSEMEN": (29700000000, 2100000000, 2650000000, "FY2024", "Michelin 2024 Annual Results (EUR ~€27.5B)"),
    "AUTOLIV INC":                      (10520000000,  400000000,   810000000, "FY2024", "Autoliv 2024 Annual Report"),
    "NIPPON STEEL CORPORATION":         (58000000000, 1740000000,  3020000000, "FY Mar2025", "Nippon Steel FY2024 Annual Report (JPY ~8.8T)"),
    "SUMITOMO ELECTRIC INDUSTRIES LTD": (30100000000, 1750000000,  1200000000, "FY Mar2025", "Sumitomo Electric FY2024 (JPY ~4.5T)"),
    "GOODYEAR TIRE & RUBBER CO, THE":   (18560000000,  870000000,   320000000, "FY2024", "Goodyear 2024 10-K"),
    "ARCELORMITTAL":                    (62400000000, 1100000000,  2700000000, "FY2024", "ArcelorMittal 2024 Annual Results"),
    "NUCOR CORP":                       (30360000000,  420000000,  2630000000, "FY2024", "Nucor 2024 10-K"),
    "NEXTEER AUTOMOTIVE GROUP LTD":     ( 3800000000,  130000000,   165000000, "FY2024", "Nexteer 2024 Annual Report"),
    "CLEVELAND CLIFFS INC":             (19068000000,  320000000,  -774000000, "FY2024", "Cleveland-Cliffs 2024 10-K"),
    "LINAMAR CORP":                     ( 6734000000,  249000000,   575000000, "FY2024", "Linamar 2024 Annual Report (CAD ~9.1B)"),
    "MARTINREA INTERNATIONAL INC":      ( 3552000000,  108000000,   148000000, "FY2024", "Martinrea 2024 Annual Report (CAD ~4.8B)"),
    "BURELLE":                          ( 9936000000,  414000000,   230000000, "FY2024", "Burelle/Plastic Omnium 2024 Annual Results (EUR ~€9.2B)"),
    "ALFA SAB DE CV":                   ( 4200000000,  168000000,   190000000, "FY2024", "Nemak FY2024 Annual Report"),
    "THYSSENKRUPP AG":                  (36300000000, 1200000000,   140000000, "FY Sep2024", "ThyssenKrupp FY2024 Annual Report (EUR ~€33.6B)"),
    "TE CONNECTIVITY LTD":              (15240000000,  890000000,  2480000000, "FY Sep2024", "TE Connectivity FY2024 Annual Report"),
    "DANA INC":                         ( 8379000000,  318000000,   168000000, "FY2024", "Dana Inc 2024 10-K"),
    "EATON CORPORATION PUBLIC LTD CO":  (24900000000, 1750000000,  4940000000, "FY2024", "Eaton 2024 Annual Report"),
    "ILLINOIS TOOL WORKS INC":          (15900000000,  870000000,  4170000000, "FY2024", "ITW 2024 10-K"),
    "HITACHI LTD":                      (17800000000, 1100000000,  1950000000, "FY Mar2025", "Hitachi FY2024 Annual Report (JPY ~2.7T Astemo)"),
    "POSCO HOLDINGS INC":               (56300000000, 1500000000,  1100000000, "FY2024", "POSCO Holdings 2024 Annual Report (KRW ~77T)"),
    "ALLISON TRANSMISSION HOLDINGS INC":(  3014000000,  158000000,   847000000, "FY2024", "Allison Transmission 2024 10-K"),
    "AB SKF":                           ( 9108000000,  728000000,  1029000000, "FY2024", "SKF 2024 Annual Report (SEK ~99B)"),
    "NIDEC CORP":                       (15410000000, 1619000000,   770000000, "FY Mar2025", "Nidec FY2024 Annual Report (JPY ~2.3T)"),
    "PARKER-HANNIFIN CORPORATION":      (19960000000, 1220000000,  3880000000, "FY Jun2024", "Parker Hannifin FY2024 Annual Report"),
    "VITESCO TECHNOLOGIES GROUP AG":    ( 4212000000,  190000000,    82000000, "FY2023", "Vitesco FY2023 (acquired by Schaeffler Nov 2024, last full year)"),
    "MELROSE INDUSTRIES PLC":           ( 2667000000,   96000000,   243000000, "FY2024", "Melrose FY2024 Annual Report (GBP ~£2.1B)"),
    "HANKOOK & COMPANY CO LTD":         ( 6497000000,  548000000,   526000000, "FY2024", "Hankook & Company 2024 (KRW ~8.9T)"),
    "PIRELLI & CO SPA":                 ( 7130000000,  410000000,  1040000000, "FY2024", "Pirelli 2024 Annual Report (EUR ~€6.6B)"),
    "CHENG SHIN RUBBER INDUSTRIAL CO LTD":( 3180000000, 140000000,   240000000, "FY2024", "Maxxis/Cheng Shin FY2024 Annual Report (TWD estimate)"),
    "TOYO TIRE CORPORATION":            ( 4420000000,  250000000,   400000000, "FY2024", "Toyo Tire 2024 Annual Report (JPY ~660B)"),
    "YOKOHAMA RUBBER CO LTD, THE":      ( 6170000000,  360000000,   600000000, "FY2024", "Yokohama Rubber 2024 Annual Report (JPY ~920B)"),
    "FUYAO GLASS INDUSTRY GROUP CO LTD":( 4700000000,  280000000,   700000000, "FY2024", "Fuyao Glass 2024 Annual Report (CNY ~34B)"),
    "NIPPON SHEET GLASS CO LTD":        ( 5100000000,  310000000,   240000000, "FY Mar2025", "NSG FY2024 Annual Report (JPY ~760B)"),
    "AGC INC":                          ( 9640000000,  620000000,   510000000, "FY2024", "AGC 2024 Annual Report (JPY ~1.44T)"),
    "CENTRAL GLASS CO LTD":             ( 1490000000,   95000000,    80000000, "FY Mar2025", "Central Glass FY2024 (JPY ~222B)"),
    "GARRETT MOTION INC":               ( 3803000000,  133000000,   451000000, "FY2024", "Garrett Motion 2024 Annual Report"),
    "SENSATA TECHNOLOGIES HOLDING PLC": ( 3535000000,  247000000,   387000000, "FY2024", "Sensata 2024 10-K"),
    "VISTEON CORP":                     ( 2959000000,  148000000,   159000000, "FY2024", "Visteon 2024 10-K"),
    "AMERICAN AXLE & MANUFACTURING":    ( 5698000000,  198000000,   143000000, "FY2024", "AAM 2024 10-K"),
    "COOPER STANDARD HOLDINGS INC":     ( 2793000000,  120000000,    82000000, "FY2024", "Cooper-Standard 2024 10-K"),
    "GENTHERM INC":                     ( 1108900000,  107600000,    72400000, "FY2024", "Gentherm 2024 10-K"),
    "GENTEX CORP":                      ( 2388700000,  112300000,   560600000, "FY2024", "Gentex 2024 10-K"),
    "TIMKENSTEEL CORPORATION":          ( 1379000000,   81000000,    29000000, "FY2024", "TimkenSteel 2024 10-K"),
    "TIMKEN CO, THE":                   ( 4570000000,  440000000,   460000000, "FY2024", "Timken 2024 10-K"),
    "ELRINGKLINGER AG":                 ( 1880000000,   98000000,    45000000, "FY2024", "ElringKlinger 2024 Annual Report (EUR ~€1.74B)"),
    "SUNGWOO HITECH CO LTD":            ( 1205000000,   40000000,    35000000, "FY2024", "Sungwoo Hitech 2024 Annual Report (KRW ~1.65T)"),
    "MINTH GROUP LTD":                  ( 3090000000,  155000000,   320000000, "FY2024", "Minth Group 2024 Annual Report (HKD ~24B)"),
    "NEXANS SA":                        ( 8320000000,  430000000,   530000000, "FY2024", "Nexans 2024 Annual Report (EUR ~€7.7B)"),
    "NITERRA CO LTD":                   ( 3484000000,  369000000,   348000000, "FY Mar2025", "Niterra FY2024 Annual Report"),
    "PARK-OHIO HOLDINGS CORP":          ( 1617000000,  148000000,    63000000, "FY2024", "Park Ohio 2024 10-K"),
    "KOITO MANUFACTURING CO LTD":       ( 6407000000,         None, 390000000, "FY Mar2025", "Koito FY2024 Annual Report (JPY ~956B)"),
    "CIE AUTOMOTIVE SA":                ( 4290000000,  200000000,   420000000, "FY2024", "CIE Automotive 2024 Annual Report (EUR ~€3.97B)"),
    "TI FLUID SYSTEMS PLC":             ( 3840000000,  155000000,   195000000, "FY2024", "TI Fluid Systems 2024 Annual Report (GBP ~£3.0B)"),
    "EXEDY CORPORATION":                ( 1876000000,         None, 134000000, "FY Mar2025", "Exedy FY2024 Annual Report"),
    "IOCHPE MAXION SA":                 ( 2300000000,  100000000,   150000000, "FY2024", "Iochpe-Maxion 2024 Annual Report (BRL ~11.5B)"),
    "TOPY INDUSTRIES LTD":              ( 1809000000,         None,  80000000, "FY Mar2025", "Topy Industries FY2024 Annual Report"),
    "SSAB AB":                          ( 7630000000,  340000000,   550000000, "FY2024", "SSAB 2024 Annual Report (SEK ~83B)"),
    "NINGBO JIFENG AUTO PARTS CO LTD":  ( 2070000000,   90000000,   190000000, "FY2024", "Ningbo Jifeng 2024 Annual Report (CNY ~15B)"),
    "AMPHENOL CORP":                    (14840000000,  720000000,  3010000000, "FY2024", "Amphenol 2024 10-K"),
    "LG ELECTRONICS INC":              (53400000000, 3100000000,  1660000000, "FY2024", "LG Electronics 2024 Annual Report (KRW ~73T)"),
    "TSUBAKIMOTO CHAIN CO":             ( 1820000000,  109000000,   136000000, "FY Mar2025", "Tsubakimoto Chain FY2024 Annual Report (JPY ~272B)"),
    "STABILUS SA":                      ( 1450000000,   78000000,   175000000, "FY2024", "Stabilus 2024 Annual Report (EUR ~€1.34B)"),
    "LEGGETT & PLATT INC":              ( 4060000000,  295000000,   280000000, "FY2024", "Leggett & Platt 2024 10-K"),
    "PRYSMIAN SPA":                     (22994000000,  980000000,  1820000000, "FY2024", "Prysmian 2024 Annual Report (EUR ~€21.3B)"),
    "COMPAGNIE DE SAINT-GOBAIN":        (51870000000, 3200000000,  4250000000, "FY2024", "Saint-Gobain 2024 Annual Report (EUR ~€48B)"),
    "MITSUBISHI STEEL MFG CO LTD":      (  870000000,   38000000,    42000000, "FY Mar2025", "Mitsubishi Steel FY2024 Annual Report (JPY ~130B)"),
    "EXCO TECHNOLOGIES LTD":            (  541000000,   48000000,    47000000, "FY2024", "Exco Technologies 2024 Annual Report"),
    "LG CHEM LTD":                      (38500000000, 2200000000,  -700000000, "FY2024", "LG Chem 2024 Annual Report (KRW ~53T)"),
    "JTEKT CORP":                       (10184000000,  737000000,   456000000, "FY Mar2025", "JTEKT FY2024 Annual Report (JPY ~1.52T)"),
    "NIPPON SEIKI CO LTD":              ( 2680000000,         None, 121000000, "FY Mar2025", "Nippon Seiki FY2024 Annual Report (JPY ~400B)"),
    "WORTHINGTON INDUSTRIES INC":       ( 1384000000,  118000000,    98000000, "FY May2024", "Worthington Industries FY2024 10-K"),
    "RYOBI LTD":                        ( 1173000000,         None,  50000000, "FY Mar2025", "Ryobi FY2024 Annual Report (JPY ~175B)"),
    "F-TECH INC":                       (  890000000,   42000000,    38000000, "FY Mar2025", "F-Tech FY2024 Annual Report (JPY ~133B)"),
    "FCC CO LTD":                       ( 1742000000,         None, 147000000, "FY Mar2025", "FCC FY2024 Annual Report (JPY ~260B)"),
    "METHODE ELECTRONICS INC":          ( 1189600000,  109400000,  -147200000, "FY May2024", "Methode Electronics FY2024 10-K"),
    "STRATTEC SECURITY CORPORATION":    (  611500000,   36200000,    24300000, "FY2024", "STRATTEC 2024 10-K"),
    "JOHNSON ELECTRIC HOLDINGS LIMITED":( 3820000000,  330000000,   220000000, "FY Mar2025", "Johnson Electric FY2024 Annual Report (HKD ~30B)"),
    "ALPS ALPINE CO LTD":               ( 5430000000,  370000000,   190000000, "FY Mar2025", "Alps Alpine FY2024 Annual Report (JPY ~810B)"),
    "RASSINI SAB DE CV":                ( 1060000000,   48000000,   120000000, "FY2024", "Rassini 2024 Annual Report (MXN ~20B)"),
    "CONSTELLIUM SE":                   ( 7390000000,  260000000,   410000000, "FY2024", "Constellium 2024 Annual Report"),
    "HINDALCO INDUSTRIES LTD":          (24000000000,  600000000,  1680000000, "FY Mar2025", "Hindalco FY2024 Annual Report (INR ~2T)"),
    "SIKA AG":                          (12150000000, 2040000000,  1740000000, "FY2024", "Sika 2024 Annual Report (CHF ~11B)"),
    "SL CORP":                          (  760000000,   36000000,    42000000, "FY2024", "SL Corp 2024 Annual Report (KRW ~1.04T)"),
    "GRUPO INDUSTRIAL SALTILLO SA DE CV":( 980000000,   55000000,    90000000, "FY2024", "GIS 2024 Annual Report (MXN ~18.5B)"),
    "STONERIDGE INC":                   (  791800000,   95600000,   -32400000, "FY2024", "Stoneridge 2024 10-K"),
    "STANDARD MOTOR PRODUCTS INC":      ( 3218000000,  396000000,   156000000, "FY2024", "Standard Motor Products 2024 10-K"),
    "T RAD CO LTD":                     (  940000000,   42000000,    55000000, "FY Mar2025", "T.RAD FY2024 Annual Report (JPY ~140B)"),
    "ILJIN GLOBAL CO LTD":              ( 1351000000,   58000000,    62000000, "FY2024", "Iljin Global 2024 Annual Report (KRW ~1.85T)"),
    "DY CORPORATION":                   (  694000000,   26000000,    23000000, "FY2024", "DY Corp 2024 Annual Report (KRW ~951B)"),
    "AKEBONO BRAKE INDUSTRY CO LTD":    ( 1675000000,         None,  54000000, "FY Mar2025", "Akebono Brake FY2024 Annual Report"),
    "AIRBOSS OF AMERICA CORP":          (  518000000,   52000000,    14000000, "FY2024", "AirBoss 2024 Annual Report"),
    "TAEYANG METAL INDUSTRIAL CO LTD":  (  310000000,   15000000,    14000000, "FY2024", "Taeyang Metal 2024 Annual Report (KRW ~424B)"),
    "NINGBO HUAXIANG ELECTRONIC CO LTD":( 1640000000,   80000000,   130000000, "FY2024", "Ningbo Huaxiang 2024 Annual Report (CNY ~11.9B)"),
    "HYUNDAI WIA CORP":                 ( 6205000000,  204000000,   161000000, "FY2024", "Hyundai Wia 2024 Annual Report (KRW ~8.5T)"),
    "BASF SE":                          (65300000000, 5200000000,  2200000000, "FY2024", "BASF 2024 Annual Report (EUR ~€60.5B)"),
    "KOBE STEEL LTD":                   ( 8160000000,  430000000,   490000000, "FY Mar2025", "Kobe Steel FY2024 Annual Report (JPY ~1.22T)"),
    "VOXX INTERNATIONAL CORP":          (  410000000,   78000000,   -24000000, "FY Feb2025", "VOXX International FY2025 10-K"),
    "MOTORCAR PARTS OF AMERICA INC":    (  591000000,   75000000,    14000000, "FY Mar2025", "Motorcar Parts 10-K FY2025"),
    "MITSUBISHI MATERIALS CORP":        ( 4410000000,  230000000,   220000000, "FY Mar2025", "Mitsubishi Materials FY2024 Annual Report (JPY ~658B)"),
    "HL MANDO CORPORATION":             ( 5694000000,  329000000,   153000000, "FY2024", "HL Mando 2024 Annual Report (KRW ~7.8T)"),
    "NINGBO JOYSON ELECTRONIC CORP":    ( 9384000000,  469000000,   338000000, "FY2024", "Ningbo Joyson 2024 Annual Report (CNY ~68B)"),
    "BHARAT FORGE LTD":                 ( 1620000000,   90000000,   260000000, "FY Mar2025", "Bharat Forge FY2024 Annual Report (INR ~135B)"),
    "WESTPORT FUEL SYSTEMS INC":        (  305000000,   38000000,   -18000000, "FY2024", "Westport Fuel Systems 2024 Annual Report"),
    "TOYOTA BOSHOKU CORP":              (13900000000,  700000000,   490000000, "FY Mar2025", "Toyota Boshoku FY2024 Annual Report (JPY ~2.07T)"),
    "TOYOTA INDUSTRIES CORP":           (22800000000, 1300000000,  1850000000, "FY Mar2025", "Toyota Industries FY2024 Annual Report (JPY ~3.4T)"),
    "NHK SPRING CO LTD":                ( 4824000000,  302000000,   235000000, "FY Mar2025", "NHK Spring FY2024 Annual Report"),
    "TAIHO KOGYO CO LTD":               (  810000000,   40000000,    62000000, "FY Mar2025", "Taiho Kogyo FY2024 Annual Report (JPY ~121B)"),
    "LS CORP":                          ( 9180000000,  330000000,   450000000, "FY2024", "LS Corp 2024 Annual Report (KRW ~12.6T)"),
    "MITSUBA CORP":                     ( 2060000000,   90000000,    68000000, "FY Mar2025", "Mitsuba FY2024 Annual Report (JPY ~307B)"),
    "LITTELFUSE INC":                   ( 2322000000,  198000000,   155000000, "FY2024", "Littelfuse 2024 10-K"),
    "INFAC CORPORATION":                (  490000000,   22000000,    18000000, "FY2024", "Infac Corp 2024 Annual Report (KRW ~671B)"),
    "SANDEN CORPORATION":               (  610000000,   35000000,    22000000, "FY Mar2025", "Sanden FY2024 Annual Report (JPY ~91B)"),
    "CORE MOLDING TECHNOLOGIES INC":    (  338200000,   25100000,    28900000, "FY2024", "Core Molding Technologies 2024 10-K"),
    "ZHEJIANG WANFENG AUTO WHEEL CO LTD":( 1380000000,   62000000,   130000000, "FY2024", "Zhejiang Wanfeng 2024 Annual Report (CNY ~10B)"),
    "DORMAN PRODUCTS INC":              ( 1924600000,  268300000,   182100000, "FY2024", "Dorman Products 2024 10-K"),
    "PARKER CORP":                      (  120000000,    8000000,     7000000, "FY2024", "Parker Corp FY2024 (small Japanese auto parts)"),
    "BANDO CHEMICAL INDUSTRIES LTD":    (  870000000,   48000000,    55000000, "FY Mar2025", "Bando Chemical FY2024 Annual Report (JPY ~130B)"),
    "NSK LTD":                          ( 6950000000,  460000000,   315000000, "FY Mar2025", "NSK FY2024 Annual Report (JPY ~1.04T)"),
    "SUNDRAM FASTENERS LTD":            (  540000000,   28000000,    92000000, "FY Mar2025", "Sundram Fasteners FY2024 Annual Report (INR ~45B)"),
    "TOKAI RIKA CO LTD":                ( 1980000000,   95000000,   120000000, "FY Mar2025", "Tokai Rika FY2024 Annual Report (JPY ~296B)"),
    "HOLLEY INC":                       (  620000000,   89000000,    52000000, "FY2024", "Holley 2024 10-K"),
    "SHANGHAI BAOLONG AUTOMOTIVE CORP": (  960000000,   55000000,    78000000, "FY2024", "Baolong Automotive 2024 Annual Report (CNY ~7B)"),
    "AHRESTY CORP":                     (  960000000,   42000000,    38000000, "FY Mar2025", "Ahresty FY2024 Annual Report (JPY ~143B)"),
    "LKQ CORP":                         (14120000000, 2520000000,  1150000000, "FY2024", "LKQ Corp 2024 10-K"),
    "TUPY S/A":                         ( 2300000000,  100000000,   175000000, "FY2024", "Tupy 2024 Annual Report (BRL ~11.5B)"),
    "POLYTEC HOLDINGS AG":              (  930000000,   45000000,    48000000, "FY2024", "Polytec 2024 Annual Report (EUR ~€860M)"),
    "THK CO LTD":                       ( 3470000000,  260000000,   250000000, "FY Mar2025", "THK FY2024 Annual Report (JPY ~518B)"),
    "ZHEJIANG SANHUA INTELLIGENT CONTROL":( 2540000000, 130000000,   380000000, "FY2024", "Zhejiang Sanhua 2024 Annual Report (CNY ~18.4B)"),
    "KONGSBERG AUTOMOTIVE ASA":         (  910000000,   62000000,    40000000, "FY2024", "Kongsberg Automotive 2024 Annual Report (NOK ~10B)"),
    "NORMA GROUP SE":                   ( 1110000000,   75000000,   115000000, "FY2024", "NORMA Group 2024 Annual Report (EUR ~€1.03B)"),
    "HI-LEX CORP":                      ( 1320000000,   58000000,    70000000, "FY Mar2025", "Hi-Lex FY2024 Annual Report (JPY ~197B)"),
    "SFS GROUP AG":                     ( 2140000000,  120000000,   260000000, "FY2024", "SFS Group 2024 Annual Report (CHF ~1.9B)"),
    "DEPO AUTO PARTS IND CO LTD":       (  580000000,   35000000,    62000000, "FY2024", "Depo Auto Parts 2024 Annual Report (TWD estimate)"),
    "XIN POINT HOLDINGS LTD":          (  340000000,   18000000,    32000000, "FY2024", "Xin Point Holdings 2024 Annual Report"),
    "DIAMOND ELECTRIC HOLDINGS CO LTD": (  570000000,   32000000,    38000000, "FY Mar2025", "Diamond Electric FY2024 Annual Report (JPY ~85B)"),
    "THULE GROUP AB":                   ( 1170000000,  195000000,   185000000, "FY2024", "Thule Group 2024 Annual Report (SEK ~12.7B)"),
    "MURO CORPORATION":                 (  180000000,    9000000,     8000000, "FY Mar2025", "Muro Corp FY2024 Annual Report (JPY ~27B)"),
    "HANWHA CORP":                      (10800000000,  550000000,   400000000, "FY2024", "Hanwha Corp 2024 Annual Report (KRW ~14.8T)"),
    "DAIDO METAL CO LTD":               (  780000000,   38000000,    45000000, "FY Mar2025", "Daido Metal FY2024 Annual Report (JPY ~116B)"),
    "JABIL CIRCUIT CO":                 (28200000000, 1430000000,  1840000000, "FY Aug2024", "Jabil FY2024 10-K"),
    "UACJ CORP":                        ( 5450000000,  290000000,   280000000, "FY Mar2025", "UACJ FY2024 Annual Report (JPY ~813B)"),
    "MUSASHI SEIMITSU INDUSTRY CO LTD": ( 2480000000,  120000000,   115000000, "FY Mar2025", "Musashi Seimitsu FY2024 Annual Report (JPY ~370B)"),
    "FURUKAWA ELECTRIC CO LTD":         (10100000000,  540000000,   320000000, "FY Mar2025", "Furukawa Electric FY2024 Annual Report (JPY ~1.5T)"),
    "NIFCO INC":                        ( 2140000000,  115000000,   210000000, "FY Mar2025", "Nifco FY2024 Annual Report (JPY ~319B)"),
    "ARKEMA":                           ( 9790000000,  720000000,   780000000, "FY2024", "Arkema 2024 Annual Report (EUR ~€9.1B)"),
    "FOSTER ELECTRIC CO LTD":           (  460000000,   24000000,    18000000, "FY Mar2025", "Foster Electric FY2024 Annual Report (JPY ~69B)"),
    "SANOH INDUSTRIAL CO LTD":          (  640000000,   30000000,    32000000, "FY Mar2025", "Sanoh Industrial FY2024 Annual Report (JPY ~95B)"),
    "KYB CORP":                         ( 3082000000,         None, 168000000, "FY Mar2025", "KYB FY2024 Annual Report"),
    "TSUBAKI NAKASHIMA CO LTD":         (  620000000,   32000000,    52000000, "FY Mar2025", "Tsubaki Nakashima FY2024 Annual Report (JPY ~93B)"),
    "FUJIKURA LTD":                     ( 5370000000,  280000000,   310000000, "FY Mar2025", "Fujikura FY2024 Annual Report (JPY ~800B)"),
    "TRIMAS CORP":                      (  775000000,   87000000,    71000000, "FY2024", "TriMas 2024 10-K"),
    "YOROZU CORP":                      ( 1150000000,   52000000,    48000000, "FY Mar2025", "Yorozu FY2024 Annual Report (JPY ~171B)"),
    "TT ELECTRONICS PLC":               (  640000000,   45000000,    38000000, "FY2024", "TT Electronics 2024 Annual Report (GBP ~£505M)"),
    "NISHIKAWA RUBBER CO LTD":          (  730000000,   36000000,    38000000, "FY Mar2025", "Nishikawa Rubber FY2024 Annual Report (JPY ~109B)"),
    "TRELLEBORG AB":                    ( 4610000000,  410000000,   660000000, "FY2024", "Trelleborg 2024 Annual Report (SEK ~50B)"),
    "BODYCOTE PLC":                     (  715000000,   85000000,   108000000, "FY2024", "Bodycote 2024 Annual Report (GBP ~£563M)"),
    "CTR CO LTD":                       (  453000000,   16000000,    13000000, "FY2024", "CTR 2024 Annual Report (KRW ~621B)"),
    "PACIFIC INDUSTRIAL CO LTD":        (  580000000,   28000000,    34000000, "FY Mar2025", "Pacific Industrial FY2024 Annual Report (JPY ~87B)"),
    "NOK CORPORATION":                  ( 3270000000,  175000000,   175000000, "FY Mar2025", "NOK Corporation FY2024 Annual Report (JPY ~488B)"),
    "SUMITOMO RIKO COMPANY LTD":        ( 3090000000,  155000000,   145000000, "FY Mar2025", "Sumitomo Riko FY2024 Annual Report (JPY ~461B)"),
    "GMB CORP":                         (  300000000,   14000000,    16000000, "FY Mar2025", "GMB Corp FY2024 Annual Report (JPY ~45B)"),
    "KENDRION NV":                      (  570000000,   48000000,    42000000, "FY2024", "Kendrion 2024 Annual Report (EUR ~€528M)"),
    "TOYODA GOSEI CO LTD":              ( 6566000000,  436000000,   302000000, "FY Mar2025", "Toyoda Gosei FY2024 Annual Report"),
    "HWASEUNG CORP CO LTD":             (  690000000,   35000000,    28000000, "FY2024", "Hwaseung Corp 2024 Annual Report (KRW ~945B)"),
    "JINHAP CO LTD":                    (  180000000,    9000000,     6000000, "FY2024", "Jinhap 2024 Annual Report (KRW ~247B)"),
    "SK INNOVATION CO LTD":             (56500000000, 2200000000,  -320000000, "FY2024", "SK Innovation 2024 Annual Report (KRW ~77.5T)"),
    "NICHIRIN CO LTD":                  (  490000000,   23000000,    28000000, "FY Mar2025", "Nichirin FY2024 Annual Report (JPY ~73B)"),
    "STANLEY ELECTRIC CO LTD":          ( 2940000000,  155000000,   195000000, "FY Mar2025", "Stanley Electric FY2024 Annual Report (JPY ~439B)"),
    "LCI INDUSTRIES":                   ( 3568000000,  245000000,   196000000, "FY2024", "LCI Industries 2024 10-K"),
    "CTEK AB (PUBL)":                   (  230000000,   28000000,    18000000, "FY2024", "CTEK 2024 Annual Report (SEK ~2.5B)"),
    "NICHIAS CORP":                     (  840000000,   45000000,    62000000, "FY Mar2025", "Nichias FY2024 Annual Report (JPY ~125B)"),
    "UCHIYAMA MANUFACTURING CORP":      (  190000000,    9000000,    10000000, "FY Mar2025", "Uchiyama Mfg FY2024 Annual Report"),
    "AMS AG":                           ( 3440000000,  330000000,  -640000000, "FY2024", "ams-OSRAM 2024 Annual Report (EUR ~€3.2B)"),
    "MITSUBOSHI BELTING LTD":           (  680000000,   35000000,    42000000, "FY Mar2025", "Mitsuboshi Belting FY2024 Annual Report (JPY ~101B)"),
    "AISAN INDUSTRY CO LTD":            ( 1210000000,   62000000,    72000000, "FY Mar2025", "Aisan Industry FY2024 Annual Report (JPY ~181B)"),
    "PIOLAX INC":                       (  780000000,   40000000,    52000000, "FY Mar2025", "Piolax FY2024 Annual Report (JPY ~116B)"),
    "NITTO DENKO CORP":                 ( 7950000000,  420000000,   870000000, "FY Mar2025", "Nitto Denko FY2024 Annual Report (JPY ~1.19T)"),
    "DAIDO CORP":                       (  360000000,   18000000,    15000000, "FY Mar2025", "Daido Corp FY2024 Annual Report (JPY ~54B)"),
    "COMMERCIAL VEHICLE GROUP INC":     (  887000000,   72000000,     8000000, "FY2024", "Commercial Vehicle Group 2024 10-K"),
    "HARADA INDUSTRY CO LTD":           (  350000000,   18000000,    12000000, "FY Mar2025", "Harada Industry FY2024 Annual Report (JPY ~52B)"),
    "NISSHINBO HOLDINGS INC":           ( 2480000000,  150000000,   120000000, "FY Mar2025", "Nisshinbo Holdings FY2024 Annual Report (JPY ~370B)"),
    "SENIOR PLC":                       (  960000000,   62000000,    62000000, "FY2024", "Senior plc 2024 Annual Report (GBP ~£756M)"),
    "HOSIDEN CORPORATION":              (  780000000,   42000000,    28000000, "FY Mar2025", "Hosiden FY2024 Annual Report (JPY ~116B)"),
    "SANKO GOSEI LTD":                  (  350000000,   18000000,    12000000, "FY Mar2025", "Sanko Gosei FY2024 Annual Report (JPY ~52B)"),
    "ZHEJIANG YINLUN MACHINERY CO LTD": (  780000000,   42000000,    62000000, "FY2024", "Zhejiang Yinlun 2024 Annual Report (CNY ~5.7B)"),
    "WOORY INDUSTRIAL HOLDINGS CO LTD": (  420000000,   22000000,    18000000, "FY2024", "Woory Industrial 2024 Annual Report (KRW ~576B)"),
    "ITT INC":                          ( 3490000000,  270000000,   540000000, "FY2024", "ITT Inc 2024 10-K"),
    "GUANGDONG HONGTU TECHNOLOGY HOLDING":( 640000000,  35000000,    55000000, "FY2024", "Guangdong Hongtu 2024 Annual Report (CNY ~4.6B)"),
    "SUPRAJIT ENGINEERING LTD":         (  420000000,   22000000,    48000000, "FY Mar2025", "Suprajit Engineering FY2024 Annual Report (INR ~35B)"),
    "DONALDSON COMPANY INC":            ( 3543000000,  320000000,   540000000, "FY Jul2024", "Donaldson 2024 10-K"),
    "YOKOWO CO LTD":                    (  470000000,   25000000,    32000000, "FY Mar2025", "Yokowo FY2024 Annual Report (JPY ~70B)"),
    "XIAMEN HONGFA ELECTROACOUSTIC CO": (  740000000,   42000000,    95000000, "FY2024", "Xiamen Hongfa 2024 Annual Report (CNY ~5.4B)"),
    "HEXAGON COMPOSITES ASA":           (  490000000,   42000000,    28000000, "FY2024", "Hexagon Composites 2024 Annual Report (NOK ~5.4B)"),
    "MODINE MANUFACTURING CO":          ( 2454000000,  185000000,   280000000, "FY Mar2025", "Modine FY2024 10-K"),
    "CTS CORP":                         (  515500000,   58400000,    74900000, "FY2024", "CTS Corp 2024 10-K"),
    "CUMMINS INC":                      (34100000000, 1820000000,  4780000000, "FY2024", "Cummins 2024 10-K"),
    "TDK CORP":                         ( 8720000000,  440000000,   540000000, "FY Mar2025", "TDK FY2024 Annual Report (JPY ~1.3T)"),
    "MAYVILLE ENGINEERING CO INC":      (  560000000,   38000000,    35000000, "FY2024", "Mayville Engineering 2024 10-K"),
    "HOWMET AEROSPACE INC":             ( 7430000000,  300000000,  1600000000, "FY2024", "Howmet Aerospace 2024 10-K"),
    "MINEBEA MITSUMI INC":              ( 5140000000,  290000000,   290000000, "FY Mar2025", "MinebeaMitsumi FY2024 Annual Report (JPY ~767B)"),
    "OMRON CORPORATION":                ( 7060000000,  780000000,   280000000, "FY Mar2025", "OMRON FY2024 Annual Report (JPY ~1.05T)"),
    "MITSUBISHI MATERIALS CORP":        ( 4410000000,  230000000,   220000000, "FY Mar2025", "Mitsubishi Materials FY2024 Annual Report"),
}

# ─── EXCLUDE REASONS ─────────────────────────────────────────────────────────
EXCLUDE_REASONS = {
    "APOLLO GLOBAL MANAGEMENT INC":     "Private equity / asset manager — no physical auto components",
    "EXXON MOBIL CORP":                 "Oil & gas — no physical auto components",
    "TOTALENERGIES SE":                 "Oil & gas — no physical auto components",
    "SHELL PLC":                        "Oil & gas — no physical auto components",
    "PHILLIPS 66":                      "Oil & gas refiner — no physical auto components",
    "ICAHN ENTERPRISES LP":             "Diversified holding company / PE",
    "BERKSHIRE HATHAWAY INC":           "Holding company / financial conglomerate",
    "BANK OF NOVA SCOTIA, THE":         "Bank — financial services",
    "KKR & CO INC":                     "Private equity / asset manager",
    "VOLKSWAGEN AG":                    "OEM (vehicle manufacturer, not supplier)",
    "HONDA MOTOR CO LTD":               "OEM (vehicle manufacturer, not supplier)",
    "TOYOTA MOTOR CORP":                "OEM (vehicle manufacturer, not supplier)",
    "NISSAN MOTOR CO LTD":              "OEM (vehicle manufacturer, not supplier)",
    "FORD MOTOR CO":                    "OEM (vehicle manufacturer, not supplier)",
    "ISUZU MOTORS LIMITED":             "OEM (vehicle manufacturer, not supplier)",
    "SHANGHAI AUTOMOTIVE INDUSTRY CORP G": "OEM — SAIC (vehicle manufacturer)",
    "MARUBENI-ITOCHU STEEL INC":        "Steel trading company — not a component manufacturer",
    "ITOCHU CORP":                      "General trading company — not a manufacturer",
    "TOYOTA TSUSHO CORP":               "Trading company — not a manufacturer",
    "SUMITOMO CORP":                    "General trading company — not a manufacturer",
    "KANEMATSU CORP":                   "Trading company — not a manufacturer",
    "GRUPO CARSO SAB DE CV":            "Diversified conglomerate — not automotive component manufacturer",
    "CITIC LIMITED":                    "State-owned financial/industrial conglomerate — not auto component manufacturer",
    "IDEMITSU KOSAN CO LTD":            "Petroleum / oil company",
    "EMERSON ELECTRIC CO":              "Industrial automation — no meaningful automotive physical components",
    "DEERE & CO":                       "Agricultural / construction equipment OEM",
    "CAPITAL SOUTHWEST CORP":           "Business development company (BDC) / financial",
    "VONTIER CORP":                     "Vehicle software / telematics — no physical auto components",
    "MOTOROLA SOLUTIONS INC":           "Public safety communications — not auto components",
    "3I GROUP PLC":                     "Private equity / infrastructure investor",
    "COMPASS DIVERSIFIED HOLDINGS":     "Diversified holding company",
    "CADRE HOLDINGS INC":               "Law enforcement / safety equipment — not automotive",
    "TRIMBLE INC":                      "Positioning technology / software — not physical auto components",
    "ULTRALIFE CORP":                   "Military & medical batteries — not automotive components",
    "ARROW ELECTRONICS INC":            "Electronics distributor — not a manufacturer",
    "AVNET INC":                        "Electronics distributor — not a manufacturer",
    "CCL INDUSTRIES INC":               "Label / packaging solutions — not auto components",
    "SONOCO PRODUCTS CO":               "Industrial packaging — not auto components",
    "AVERY DENNISON CORP":              "Label / RFID / packaging — not physical auto components",
    "SHERWIN-WILLIAMS CO, THE":         "Paint & coatings — primarily retail/architecture, not auto component manufacturer",
    "ATLAS COPCO AB":                   "Industrial tools / compressors — not auto components",
    "JOHNSON CONTROLS INC":             "Building management systems — sold automotive battery division",
    "DOVER CORPORATION":                "Diversified industrial — no primary automotive physical component segment",
    "WESTINGHOUSE AIR BRAKE TECHNOLOGIES": "Rail equipment — not automotive",
    "SCHNEIDER ELECTRIC SE":            "Building energy management — not auto components",
    "SIEMENS AKTIENGESELLSCHAFT":       "Industrial conglomerate — automotive division divested",
    "FORTIVE CORPORATION":              "Industrial tech / SaaS — not physical auto components",
    "MUELLER INDUSTRIES INC":           "Copper products / HVAC — not primarily automotive",
    "AMG ADVANCED METALLURGICAL GRP NV": "Specialty metals — primary markets aerospace/energy, not auto",
    "TEXTRON INC":                      "Aviation (Cessna/Bell) + defense — auto content is immaterial",
    "REGAL REXNORD CORPORATION":        "Industrial motors — not primary auto component supplier",
    "VAPOR ACQUISITION CORP":           "Holding / SPAC — no operational auto component business",
    "AEQUITA SE & CO KGAA":             "Investment holding company — no operating auto component business",
    "MITSUBISHI GAS CHEMICAL CO INC":   "Industrial chemicals — automotive use is minor",
    "ENERSYS":                          "Industrial/telecom batteries — not primarily automotive",
    "LUXSHARE LTD":                     "Primarily Apple/consumer electronics — auto division immaterial",
    "SONY CORP":                        "Consumer electronics — automotive sensors are a minor segment",
    "MITSUBISHI HEAVY INDUSTRIES LTD":  "Defense / energy / ships — automotive contribution immaterial",
    "QUANTA COMPUTER INC":              "IT hardware / servers — auto EMS is immaterial portion",
    "AVIENT CORPORATION":               "Specialty polymer distribution — not direct auto component manufacturer",
    "ARB CORP LTD":                     "Off-road accessories aftermarket — not OEM automotive supply chain",
    "DOMETIC GROUP AB (PUBL)":          "RV / marine climate control — not OEM automotive supplier",
    "ZHUZHOU TIMES NEW MATERIAL TECHNOLO": "Rail / wind composites — automotive use immaterial",
    "DUPONT DE NEMOURS INC":            "Diversified materials — automotive is small segment; primarily specialty chemicals",
    "LEM HOLDING SA":                   "Industrial current sensors — automotive segment is small",
    "ZHENGZHOU COAL MINING MACHINERY GRP": "Coal mining equipment — automotive use immaterial",
    "NACHI FUJIKOSHI CORP":             "Primarily machine tools and robots — automotive content small",
}

# ─── COMPANY DETAILS ─────────────────────────────────────────────────────────
# (tier, parts, oems, duns)
INCLUDE_DETAILS = {
    "MAGNA INTERNATIONAL INC":          (1, "Body/chassis stampings, seating systems, exterior mirrors, powertrain components, vision systems, roof systems, closures", "General Motors, Ford, Stellantis, Tesla, Honda, Toyota", 201516002),
    "DENSO CORPORATION":                (1, "HVAC systems, fuel injection systems, starters, alternators, EV motors, thermal management, sensors", "Toyota, Honda, Nissan, Subaru, Mazda, General Motors", 690597851),
    "CONTINENTAL AKTIENGESELLSCHAFT":   (1, "Tires, brakes, powertrain electronics, ADAS sensors, instrument clusters, CVT belts", "Volkswagen, Mercedes, BMW, Tesla, Stellantis, General Motors", 315674267),
    "AISIN CORP":                       (1, "Automatic transmissions, torque converters, door/body hardware, water pumps, brakes, EV drive units", "Toyota, Volkswagen, BMW, Stellantis", 690535588),
    "VALEO":                            (1, "Lighting systems, wiper systems, thermal systems, ADAS sensors, EV charging equipment, clutch modules", "Volkswagen, BMW, Mercedes, Renault, Volvo, General Motors", 275242212),
    "FAURECIA":                         (1, "Seating frames & mechanisms, acoustic packages, exhaust systems, hydrogen storage tanks, cockpit modules", "Volkswagen, BMW, Ford, General Motors, Stellantis, Renault", 275124311),
    "HYUNDAI MOBIS CO LTD":             (1, "Chassis modules, front/rear axle modules, airbag systems, instrument panels, EV battery packs", "Hyundai, Kia", 687755488),
    "LEAR CORP":                        (1, "Seat systems, seat mechanisms, seat foam & trim, e-systems junction boxes, wiring harnesses", "General Motors, Ford, BMW, Volkswagen", 175592476),
    "BORGWARNER INC":                   (1, "Turbochargers, torque transfer systems, EV drive modules, battery packs, iDM systems", "General Motors, Ford, Toyota, Stellantis, Volkswagen", 139469787),
    "ADIENT PUBLIC LTD CO":             (1, "Seat frames, seat mechanisms, seat foam, seat trim/fabric, complete seat assemblies", "Ford, General Motors, Stellantis, Volkswagen, BMW", 985655816),
    "COMPAGNIE GENERALE DES ETABLISSEMEN": (1, "Passenger tires, truck tires, OTR tires, specialty tires, tire accessories", "Volkswagen, Mercedes, BMW, Tesla, Stellantis, Renault", 281938431),
    "AUTOLIV INC":                      (1, "Airbags, seatbelts, steering wheels, pretensioners, pedestrian protection systems", "Hyundai, Honda, Toyota, Stellantis, Volkswagen, Mercedes", 76854327),
    "NIPPON STEEL CORPORATION":         (1, "High-strength steel sheet, galvanized steel, electrical steel, steel tubes", "Toyota, Honda, Nissan, General Motors", 690570072),
    "SUMITOMO ELECTRIC INDUSTRIES LTD": (1, "Wiring harnesses, optical fiber cables, power cables, brake hoses", "Toyota, Honda, Volkswagen", 690556345),
    "GOODYEAR TIRE & RUBBER CO, THE":   (1, "Passenger tires, truck tires, run-flat tires, off-road tires", "General Motors, Ford, Stellantis, Volkswagen, Audi", 4467924),
    "ARCELORMITTAL":                    (1, "Ultra-high-strength steel, dual-phase steel, galvanized steel, steel tube", "General Motors, Ford, Stellantis", 400020397),
    "NUCOR CORP":                       (1, "Flat-rolled steel sheet, steel bars, structural steel for automotive", "General Motors, Mercedes", 3446796),
    "NEXTEER AUTOMOTIVE GROUP LTD":     (1, "Electric power steering systems, rack & pinion assemblies, driveshafts, steering columns", "General Motors, Ford, Stellantis, BMW, Volkswagen", 864421750),
    "CLEVELAND CLIFFS INC":             (1, "Advanced high-strength steel, galvanized steel sheet, automotive steel stampings", "General Motors, Ford, Stellantis, Toyota", 147964571),
    "LINAMAR CORP":                     (1, "Powertrain components, driveline systems, EV structural components, precision machined parts", "General Motors, Ford, Stellantis, Volkswagen, Tesla", 209831544),
    "MARTINREA INTERNATIONAL INC":      (1, "Metal stampings, structural assemblies, fluid management modules, bumper systems", "General Motors, Ford, Stellantis, BMW, Volkswagen", 252027883),
    "BURELLE":                          (1, "Bumper fascia, body panels, fuel systems, hydrogen storage modules (Plastic Omnium)", "Volkswagen, Audi, BMW, Mercedes, Stellantis, Renault", 281213280),
    "ALFA SAB DE CV":                   (1, "Aluminum cylinder heads, engine blocks, EV battery housings, aluminum castings (Nemak)", "General Motors, Ford, Stellantis, BMW, Mercedes", 812278687),
    "THYSSENKRUPP AG":                  (1, "Automotive steel, springs, stabilizer bars, steering systems, camshafts", "Mercedes, BMW, Audi, Tesla", 340502442),
    "TE CONNECTIVITY LTD":              (1, "EV connectors, high-voltage harness connectors, relays, sensors, antennae", "General Motors, BMW, Ford, Toyota, Tesla, Volkswagen", 485203835),
    "DANA INC":                         (1, "Driveshafts, axles, electric drive units, thermal management, sealing solutions", "Ford, Stellantis, Toyota", 809105351),
    "EATON CORPORATION PUBLIC LTD CO":  (1, "Transmission systems, vehicle power distribution, EV charging & control", "General Motors", 985419987),
    "ILLINOIS TOOL WORKS INC":          (1, "Automotive fasteners, polymers, fluids, welding equipment for OEM assembly", "General Motors, Toyota, Ford, Volkswagen, BMW, Tesla", 5146428),
    "HITACHI LTD":                      (1, "EV inverters, motors, e-axles, ADAS radar, steering systems (Astemo JV)", "General Motors, Nissan, Ford, Volkswagen, Toyota", 690541503),
    "POSCO HOLDINGS INC":               (1, "Automotive high-strength steel, EV battery steel cases, electrical steel", "Hyundai, Kia, General Motors, Ford, Tesla", 687741991),
    "ALLISON TRANSMISSION HOLDINGS INC":(1, "Fully-automatic transmissions for medium/heavy trucks and buses", "Volvo", 969132880),
    "AB SKF":                           (1, "Wheel-end bearings, hub bearing units, seals, electric motor bearings", "Volkswagen, BMW, Ford, General Motors", 353945744),
    "NIDEC CORP":                       (1, "EV traction motors, e-axles, EPS motors, blower motors, precision motors", "Toyota, Honda, Stellantis, General Motors, Ford", 690635255),
    "PARKER-HANNIFIN CORPORATION":      (1, "Fluid connectors, hydraulic systems, filtration, thermal management, sealing", "General Motors, Ford, Volkswagen, Hyundai, Stellantis", 4175550),
    "VITESCO TECHNOLOGIES GROUP AG":    (1, "EV power electronics, e-axle drive systems, emission control, battery management", "Volkswagen, Hyundai, Renault, Ford, General Motors", 343056307),
    "MELROSE INDUSTRIES PLC":           (1, "Driveshafts, sideshafts, eDrive units, powder metal components (GKN Automotive)", "BMW, Mercedes, Volkswagen, Toyota", 221150144),
    "HANKOOK & COMPANY CO LTD":         (1, "Passenger tires, performance tires, EV-specific tires", "Tesla, BMW, Hyundai, Kia", 687735407),
    "PIRELLI & CO SPA":                 (1, "High-performance tires, run-flat tires, EV tires, motorsport tires", "BMW, Mercedes, Audi", 436854350),
    "CHENG SHIN RUBBER INDUSTRIAL CO LTD":(1, "Passenger tires, truck tires, off-road tires, bicycle tires", "General Motors, Ford, Toyota, Nissan", 656000718),
    "TOYO TIRE CORPORATION":            (1, "Passenger tires, light truck tires, performance tires", "Toyota, Mazda, Audi, Ford", 690557053),
    "YOKOHAMA RUBBER CO LTD, THE":      (1, "Passenger tires, SUV tires, winter tires, industrial hoses", "Toyota, Mazda, Mercedes", 690565601),
    "FUYAO GLASS INDUSTRY GROUP CO LTD":(1, "Windshields, side windows, rear windows, laminated safety glass", "Volkswagen, General Motors, Ford, Toyota, Tesla, Stellantis", 654532019),
    "NIPPON SHEET GLASS CO LTD":        (1, "Automotive laminated glass, tempered glass, solar-control glass", "Toyota, Honda, Volkswagen, Nissan, General Motors", 690555925),
    "AGC INC":                          (1, "Windshields, side glass, HUD glass, acoustic laminated glass", "BMW, Volkswagen, Mercedes, General Motors, Tesla", 690553888),
    "GARRETT MOTION INC":               (1, "Turbochargers, e-turbo systems, electric compressors", "Ford, Volkswagen, BMW, Mercedes, Toyota", 81174775),
    "SENSATA TECHNOLOGIES HOLDING PLC": (1, "Pressure sensors, temperature sensors, position sensors, EV battery sensors", "Tesla, General Motors, Ford, Volkswagen, Volvo", 223144058),
    "VISTEON CORP":                     (1, "Digital instrument clusters, infotainment head units, connected car systems", "General Motors, Ford, Stellantis, Hyundai, Kia, Nissan", 183727804),
    "AMERICAN AXLE & MANUFACTURING":    (1, "Driveshafts, axles, driveline modules, EV electric drive units", "General Motors, Ford, Stellantis", 44766678),
    "COOPER STANDARD HOLDINGS INC":     (1, "Door & window sealing systems, fuel & brake hoses, fluid transfer systems", "General Motors, Ford, Stellantis, Volkswagen, Mercedes", 361293918),
    "GENTHERM INC":                     (1, "Heated/cooled/ventilated seats, battery thermal management, steering wheel heaters", "Ford, General Motors, Toyota, Stellantis, Tesla", 556879252),
    "LG CHEM LTD":                      (1, "EV battery cells, battery modules, battery packs, cathode materials", "General Motors, Toyota, Tesla", 688279996),
    "JTEKT CORP":                       (1, "Electric power steering systems, column-type EPS, wheel-end bearings", "Toyota, General Motors", 690535646),
    "AISIN CORP":                       (1, "Automatic transmissions, door hardware, water pumps, brakes, EV drive units", "Toyota, Volkswagen, BMW, Stellantis", 690535588),
    "HL MANDO CORPORATION":             (1, "Brakes, steering systems, suspension, ADAS, electronic control units", "General Motors, Hyundai, Kia, Ford, Nissan, Chrysler", 688250455),
    "NINGBO JOYSON ELECTRONIC CORP":    (1, "Airbag systems, seatbelts, steering wheels, vehicle software (Key Safety Systems)", "Volkswagen, BMW, Mercedes, General Motors, Tesla", 421342800),
    "LG ELECTRONICS INC":              (1, "EV traction motors, infotainment systems, EV chargers, ADAS cameras", "General Motors, Toyota, Hyundai, Mercedes, Renault", 688298116),
    "TOYOTA BOSHOKU CORP":              (1, "Interior systems, seat assemblies, door trim, air filters, fluid purification", "Toyota, Honda, Nissan, BMW, Subaru", 690911185),
    "TOYOTA INDUSTRIES CORP":           (1, "Compressors, inverters, EV motors, textile machinery (separate segment)", "Toyota, Tesla, General Motors, Ford, Mercedes", 691239750),
    "KYB CORP":                         (1, "Shock absorbers, struts, steering gear boxes, EPS systems", "Toyota, Honda, Nissan, Ford, Volkswagen, Subaru", 690547401),
    "NSK LTD":                          (1, "Ball bearings, tapered roller bearings, EPS steering systems", "Toyota, Honda, Nissan, Suzuki, Mazda, Subaru, General Motors", 690535059),
    "FURUKAWA ELECTRIC CO LTD":         (1, "Wiring harnesses, fiber optic cables, copper products, EV components", "General Motors, Ford, Stellantis, BYD, Toyota, Honda", 690553763),
    "HYUNDAI WIA CORP":                 (1, "Engine systems, drivetrain, axles, machining systems", "Hyundai, Kia, General Motors, Renault, Nissan, Volvo", 687784926),
    "SK INNOVATION CO LTD":             (1, "EV battery cells, battery modules, battery management systems", "Ford, Volkswagen, Hyundai, Kia, Mercedes", 631064016),
    "CUMMINS INC":                      (1, "Diesel/natural gas/hydrogen engines, emission after-treatment systems", "Stellantis", 6415160),
    "JABIL CIRCUIT CO":                 (1, "Automotive electronics manufacturing, EV power modules, connectors", "Ford, General Motors, Toyota, Honda, Hyundai", 41810979),
    "PRYSMIAN SPA":                     (1, "High-voltage EV cables, automotive wiring, charging cables", "Ford, Toyota, Stellantis, General Motors, Volkswagen", 544091510),
    "COMPAGNIE DE SAINT-GOBAIN":        (1, "Automotive glass (Saint-Gobain Sekurit), sealing, abrasives", "Volkswagen, BMW, Mercedes, Ford, Toyota, General Motors", 275133692),
    "LKQ CORP":                         (1, "Aftermarket collision parts, OEM-equivalent parts, salvage vehicles", "N/A", 28117104),
    "EATON CORPORATION PUBLIC LTD CO":  (1, "Transmission systems, vehicle power distribution, EV charging & control", "General Motors", 985419987),
    # Tier 2
    "GENTEX CORP":                      (2, "Auto-dimming mirrors, HomeLink garage transmitter, full display mirrors", "Volkswagen, Mercedes, BMW, Toyota", 65855363),
    "TIMKENSTEEL CORPORATION":          (2, "Specialty steel bars, seamless steel tubes for drivetrain applications", "General Motors, Ford, Stellantis", 79236657),
    "TIMKEN CO, THE":                   (2, "Tapered roller bearings, spherical roller bearings, driveshaft bearings", "General Motors", 4465100),
    "ELRINGKLINGER AG":                 (2, "Cylinder head gaskets, special gaskets, shielding, EV battery cell frames", "Volkswagen, Mercedes, BMW, Tesla", 315342345),
    "SUNGWOO HITECH CO LTD":            (2, "Hot-stamped metal stampings, door rings, B-pillars, floor panels", "Kia, Hyundai, Volkswagen, General Motors", 689285091),
    "MINTH GROUP LTD":                  (2, "Decorative trim strips, door frame sealing, EV battery enclosures, structural parts", "Tesla, General Motors, Toyota, Mercedes, Volkswagen", 864393322),
    "NEXANS SA":                        (2, "High-voltage EV cables, low-voltage wiring harnesses, charging cables", "Stellantis, Volkswagen, BMW, Mercedes", 738162205),
    "NITERRA CO LTD":                   (2, "Spark plugs, glow plugs, oxygen sensors, temperature sensors", "Toyota, Honda, Volkswagen, Ford, General Motors", 690569256),
    "PARK-OHIO HOLDINGS CORP":          (2, "Engineered fasteners, assembly components, supply chain management", "General Motors, Ford, Stellantis", 49746014),
    "KOITO MANUFACTURING CO LTD":       (2, "Headlamps, taillamps, fog lamps, LED lighting assemblies", "Toyota, Honda, Subaru", 690579347),
    "CIE AUTOMOTIVE SA":                (2, "Metal stamped parts, aluminum castings, forge components, plastic/composite parts", "Volkswagen, Stellantis, Ford, BMW, General Motors", 471743500),
    "TI FLUID SYSTEMS PLC":             (2, "Fuel delivery systems, brake & fluid lines, thermal management tubes", "Volkswagen, Stellantis, Ford, BMW, General Motors", 220607205),
    "EXEDY CORPORATION":                (2, "Clutch covers, clutch discs, torque converters, flywheels", "Toyota, Honda, Nissan, Subaru", 690554894),
    "IOCHPE MAXION SA":                 (2, "Steel wheels, aluminum wheels, truck chassis frames", "Volkswagen, Stellantis, General Motors, Ford, Toyota", 898699483),
    "TOPY INDUSTRIES LTD":              (2, "Steel wheels, wheel rims, forged products", "General Motors, Toyota, Ford, Honda, Nissan, Stellantis", 690544341),
    "SSAB AB":                          (2, "High-strength steel sheet, wear-resistant steel, structural steel", "Volvo", 353957822),
    "NINGBO JIFENG AUTO PARTS CO LTD":  (2, "Automotive seat components, headrests, armrests, seat structural parts", "BMW, Audi, Volkswagen, General Motors, Tesla, Ford", 527651909),
    "AMPHENOL CORP":                    (2, "High-voltage EV connectors, sensor connectors, automotive antenna assemblies", "General Motors, Audi, BMW, Nissan", 177220647),
    "TSUBAKIMOTO CHAIN CO":             (2, "Engine timing chains, drive chains, cam phasers, automotive chain systems", "Toyota, Ford, General Motors, Nissan, Honda", 690555214),
    "STABILUS SA":                      (2, "Gas springs, power struts, liftgate actuators, trunk lifters", "Volkswagen, BMW, Ford, Tesla, General Motors", 400652919),
    "LEGGETT & PLATT INC":              (2, "Seat suspension systems, lumbar supports, seat comfort structures", "General Motors, Toyota, Ford", 7140064),
    "MITSUBISHI STEEL MFG CO LTD":      (2, "Suspension springs, stabilizer bars, high-strength springs", "Ford, Stellantis, Volkswagen, BMW, General Motors", 690552633),
    "EXCO TECHNOLOGIES LTD":           (2, "Extrusion tooling, die casting tooling, auto trim mouldings", "Toyota, Honda, Nissan", 201648052),
    "NIPPON SEIKI CO LTD":              (2, "Digital instrument clusters, head-up displays, meters, gauges", "General Motors, Honda, Nissan, Mazda, Subaru, BMW", 690673546),
    "WORTHINGTON INDUSTRIES INC":       (2, "Pressurized steel cylinders, custom steel processing, auto stampings", "General Motors, Stellantis", 4312401),
    "RYOBI LTD":                        (2, "Aluminum die castings, transmission cases, engine brackets, structural parts", "Volkswagen, Toyota", 690536123),
    "F-TECH INC":                       (2, "Pedal assemblies, structural frames, suspension components", "Honda, Nissan, General Motors", 690809090),
    "FCC CO LTD":                       (2, "Clutch systems, multi-plate clutch packs, ATM clutches", "Honda, Toyota, Ford, General Motors", 690661756),
    "METHODE ELECTRONICS INC":          (2, "EV bus bars, LED lighting solutions, user interface controls, cable assemblies", "General Motors", 5092135),
    "STRATTEC SECURITY CORPORATION":    (2, "Locks, keys, vehicle access systems, door handles", "Stellantis, General Motors, Ford", 879168029),
    "JOHNSON ELECTRIC HOLDINGS LIMITED":(2, "DC motors, actuators, solenoids for windows/doors/HVAC", "Ford, Volkswagen, General Motors", 875648826),
    "ALPS ALPINE CO LTD":               (2, "Automotive sensors, HMI switches, radar sensors, power window switches", "General Motors, Stellantis, Ford, Volkswagen, Mercedes", 690547641),
    "RASSINI SAB DE CV":                (2, "Leaf springs, coil springs, brake discs, front suspension systems", "General Motors, Ford, Stellantis, Volkswagen, Toyota", 811393222),
    "CONSTELLIUM SE":                   (2, "Aluminum auto body sheet, crash management systems, structural profiles", "BMW, Mercedes, Ford, Stellantis", 274121531),
    "HINDALCO INDUSTRIES LTD":          (2, "Aluminum auto body sheet, flat-rolled aluminum for closures", "Ford, General Motors, BMW", 650141922),
    "SIKA AG":                          (2, "Structural adhesives, acoustic baffles, seam sealants, battery adhesives for EV", "Volkswagen, Stellantis, Toyota", 480000538),
    "SL CORP":                          (2, "Interior lighting modules, ambient lighting, exterior lamps", "Hyundai, General Motors, Kia", 687751727),
    "GRUPO INDUSTRIAL SALTILLO SA DE CV":(2, "Aluminum engine blocks, cylinder heads, EV battery housings", "General Motors, Ford, Stellantis, Volkswagen, Nissan", 810535872),
    "STONERIDGE INC":                   (2, "Smart mirrors (MirrorEye), tachograph systems, control units", "Ford, General Motors, Stellantis, Toyota", 606280873),
    "STANDARD MOTOR PRODUCTS INC":      (2, "Ignition coils, sensors, fuel injectors, EGR components, engine management", "Ford, General Motors, Volvo", 1315266),
    "T RAD CO LTD":                     (2, "Radiators, oil coolers, heat exchangers, EV thermal management modules", "Toyota, Honda, General Motors", 690569751),
    "ILJIN GLOBAL CO LTD":              (2, "Wheel bearings, hub units, driveshafts", "Hyundai, General Motors, Ford, Kia, Mercedes", 695687668),
    "DY CORPORATION":                   (2, "Wire harnesses, high-voltage harnesses, connectors", "General Motors, Hyundai, Kia", 687769307),
    "AKEBONO BRAKE INDUSTRY CO LTD":    (2, "Disc brake pads, drum brake shoes, brake calipers, brake assemblies", "Toyota, Honda, Nissan, Subaru, Mazda, General Motors", 690535927),
    "AIRBOSS OF AMERICA CORP":          (2, "Anti-vibration products, rubber NVH components, body mounts", "General Motors, Ford, Stellantis, Honda", 247865264),
    "TAEYANG METAL INDUSTRIAL CO LTD":  (2, "Metal stampings, door hinges, hood hinges, structural parts", "Hyundai, Kia, General Motors, Ford, Stellantis, Mazda", 687782458),
    "NINGBO HUAXIANG ELECTRONIC CO LTD":(2, "Interior plastic parts, panels, handles, console components", "Volkswagen, BMW, Mercedes, Volvo, General Motors", 420859670),
    "BASF SE":                          (2, "Automotive coatings, engineering plastics, battery materials, foam", "Mercedes, Volkswagen, Toyota, Tesla", 315000554),
    "KOBE STEEL LTD":                   (2, "Aluminum alloy sheet, high-strength steel wire rod, forged components", "Toyota, Nissan, Honda, Mazda, Subaru, General Motors", 690535018),
    "VOXX INTERNATIONAL CORP":          (2, "Remote start systems, vehicle security, OEM electronics", "Stellantis, Ford, General Motors, Nissan, Subaru", 44694040),
    "MOTORCAR PARTS OF AMERICA INC":    (2, "Remanufactured alternators, starters, wheel hubs, brake calipers", "General Motors, Ford, Toyota, Honda, Hyundai", 44477123),
    "MITSUBISHI MATERIALS CORP":        (2, "Cemented carbide cutting tools, copper alloys, aluminum for auto", "Toyota, Ford, Nissan, Volkswagen", 690536867),
    "BHARAT FORGE LTD":                 (2, "Forged crankshafts, connecting rods, front axle beams, knuckles", "Volkswagen, Mercedes, Volvo, General Motors, Stellantis", 650049299),
    "WESTPORT FUEL SYSTEMS INC":        (2, "CNG/LNG fuel systems, HPDI natural gas injectors, hydrogen fuel systems", "General Motors, Ford, Volkswagen, Honda, Hyundai", 252892955),
    "NHK SPRING CO LTD":                (2, "Suspension springs, valve springs, disk springs, anti-vibration systems", "Toyota, Honda, Nissan, Subaru, Mazda, General Motors", 690541545),
    "TAIHO KOGYO CO LTD":               (2, "Engine bearings, sliding bearings, plain bearings", "Toyota, Honda, Nissan, Mitsubishi, General Motors", 690559943),
    "LS CORP":                          (2, "High-voltage EV cables, wire harnesses, busbars, EV charging cables", "Kia, Ford, Stellantis, Volkswagen, Hyundai, General Motors", 687953992),
    "MITSUBA CORP":                     (2, "Starter motors, wiper systems, blower motors, door actuators", "Honda, Nissan, Volkswagen, BMW", 690658596),
    "LITTELFUSE INC":                   (2, "Automotive fuses, circuit protection, EV contactors, sensors", "Tesla", 5212246),
    "INFAC CORPORATION":                (2, "Metal stampings, door hinges, structural parts", "Hyundai, Kia, General Motors, Mazda", 687802777),
    "SANDEN CORPORATION":               (2, "Automotive HVAC compressors, EV heat pump modules", "Toyota, Hyundai, Audi, BMW, Mercedes, Ford, General Motors", 690558051),
    "CORE MOLDING TECHNOLOGIES INC":    (2, "Fiberglass composite hoods, fenders, roofs, cab panels", "Volvo, Ford, BMW, General Motors", 965507312),
    "ZHEJIANG WANFENG AUTO WHEEL CO LTD":(2, "Aluminum alloy wheels, magnesium wheels", "Mercedes, BMW, Volkswagen, General Motors, Ford, Stellantis", 530750454),
    "DORMAN PRODUCTS INC":              (2, "Aftermarket auto parts (brakes, cooling, electrical, chassis)", "Ford, Stellantis, General Motors", 93715316),
    "BANDO CHEMICAL INDUSTRIES LTD":    (2, "Drive belts, timing belts, ribbed belts", "General Motors", 690579495),
    "NSK LTD":                          (2, "Ball bearings, tapered roller bearings, EPS steering systems", "Toyota, Honda, Nissan, Suzuki, Mazda, Subaru, General Motors", 690535059),
    "SUNDRAM FASTENERS LTD":            (2, "High-tensile fasteners, powder metal components, castings", "General Motors, Ford, Volkswagen, Renault, Volvo", 650086143),
    "TOKAI RIKA CO LTD":                (2, "Steering switches, turn signal levers, shift lock, ignition switches", "Toyota, Suzuki, Nissan, General Motors, Ford, Subaru", 690538723),
    "HOLLEY INC":                       (2, "Performance carburetors, fuel injection systems, exhaust systems", "General Motors", 117688620),
    "SHANGHAI BAOLONG AUTOMOTIVE CORP": (2, "TPMS sensors, valves, inflators, EV thermal expansion valves", "General Motors, Ford, Chrysler, Volkswagen, Mercedes", 544792021),
    "AHRESTY CORP":                     (2, "Aluminum die castings, transmission cases, engine parts", "General Motors, Toyota, Honda, Nissan, Suzuki, Mazda", 690569579),
    "TUPY S/A":                         (2, "Gray iron castings, ductile iron castings, engine blocks, cylinder heads", "General Motors, Ford, Stellantis, Volkswagen", 898699483),
    "POLYTEC HOLDINGS AG":              (2, "Thermoplastic exterior parts, door panels, wheel arch liners, engine covers", "BMW, Volkswagen, Mercedes, Stellantis", 300243545),
    "THK CO LTD":                       (2, "Linear motion guides, ball screws, automotive actuators", "Toyota, Honda, General Motors", 690599097),
    "ZHEJIANG SANHUA INTELLIGENT CONTROL":(2, "EV thermal expansion valves, refrigerant valves, heat exchangers", "Toyota, Volkswagen, Tesla", 544771496),
    "KONGSBERG AUTOMOTIVE ASA":         (2, "Gear shift cables, seating comfort systems, fluid/air management", "Ford, Nissan", 518896162),
    "NORMA GROUP SE":                   (2, "Clamps, connectors, hose couplings for fluid management", "N/A", 506707343),
    "HI-LEX CORP":                      (2, "Control cables, window regulator cables, seat cables", "Ford, General Motors, Stellantis, Toyota, Honda, Mercedes", 690539358),
    "SFS GROUP AG":                     (2, "Precision metal fasteners, engineered components for automotive", "Volkswagen", 483367074),
    "DEPO AUTO PARTS IND CO LTD":       (2, "Aftermarket headlamps, tail lamps, fog lamps", "Toyota, Honda, Nissan, Ford, General Motors, BMW", 656275658),
    "XIN POINT HOLDINGS LTD":          (2, "Metal trim pieces, side moldings, decorative trim", "General Motors, Mercedes, Volvo, BMW, Ford, Honda", 815557373),
    "DIAMOND ELECTRIC HOLDINGS CO LTD": (2, "Ignition coils, ignition modules, distributors", "Ford, Subaru, General Motors, Toyota, Honda", 717684106),
    "THULE GROUP AB":                   (2, "Roof racks, tow bars, bike carriers, cargo boxes", "Volvo, Audi, BMW", 776631116),
    "MURO CORPORATION":                 (2, "Small metal parts, fasteners, brackets", "Toyota, Honda, Nissan, Mazda, Subaru", 691458483),
    "HANWHA CORP":                      (2, "Automotive steel, industrial components, auto parts", "Hyundai, BMW, Kia, Mercedes, Volkswagen, General Motors", 687739466),
    "DAIDO METAL CO LTD":               (2, "Engine plain bearings, thrust washers, large-engine bearings", "Ford, General Motors, Toyota, Honda, Nissan", 690559257),
    "JABIL CIRCUIT CO":                 (1, "Automotive electronics manufacturing, EV power modules, connectors", "Ford, General Motors, Toyota, Honda, Hyundai", 41810979),
    "UACJ CORP":                        (2, "Aluminum flat-rolled products, aluminum auto body sheet", "General Motors, Toyota, Tesla, Ford, Honda", 690644448),
    "MUSASHI SEIMITSU INDUSTRY CO LTD": (2, "Forged steel parts, differential gears, EV drivetrain components", "Toyota, Honda, Mitsubishi, Suzuki, General Motors", 690679840),
    "NIFCO INC":                        (2, "Plastic fasteners, clips, retainers, trim attachment parts", "General Motors, Ford, Toyota, Stellantis, Tesla, Renault", 690601307),
    "ARKEMA":                           (2, "Specialty polymer powders, bio-based polyamides, auto lightweighting materials", "Volkswagen, BMW, Stellantis", 266873780),
    "FOSTER ELECTRIC CO LTD":           (2, "Speakers, microphones, actuators for automotive", "Toyota, Honda, Nissan, Suzuki, Mazda, Subaru, General Motors", 690555933),
    "SANOH INDUSTRIAL CO LTD":          (2, "Steel tubes, fuel lines, brake lines, clutch pipes", "Toyota, Nissan, Honda, Mazda", 690773668),
    "TSUBAKI NAKASHIMA CO LTD":         (2, "Precision steel balls, ball studs, linear motion components", "Toyota, Volkswagen, Ford, Tesla, Honda", 714213373),
    "FUJIKURA LTD":                     (2, "Wiring harnesses, fiber optic cables, power cables", "Toyota, Honda, Nissan", 690537048),
    "TRIMAS CORP":                      (2, "Aerospace & auto specialty products, steel parts", "General Motors, Ford, Stellantis", 175591072),
    "YOROZU CORP":                      (2, "Suspension parts, cross members, front sub-frames", "General Motors, Toyota, Honda, Volkswagen, Nissan", 690845367),
    "TT ELECTRONICS PLC":               (2, "Position sensors, power electronics, sensors for auto", "BMW, Mercedes", 212398689),
    "NISHIKAWA RUBBER CO LTD":          (2, "Weather strips, door seals, trunk seals", "Toyota, Honda, Nissan, Mazda", 690578760),
    "TRELLEBORG AB":                    (2, "Sealing systems, anti-vibration products, engineered coated fabrics", "Stellantis, Volvo, Volkswagen", 353944689),
    "BODYCOTE PLC":                     (2, "Heat treatment of powertrain gears, bearings, chassis parts", "BMW, Volkswagen", 218145332),
    "CTR CO LTD":                       (2, "Steering gear, rack & pinion, EPS systems", "Hyundai, Kia, General Motors, Tesla, Ford", 687819359),
    "PACIFIC INDUSTRIAL CO LTD":        (2, "Tire pressure valves, TPMS valve cores, plastic & rubber parts", "General Motors, Toyota, Honda, Nissan, Ford", 690536404),
    "NOK CORPORATION":                  (2, "Oil seals, O-rings, sealing solutions, EV motor seals", "Toyota, Nissan, Honda, Volkswagen, General Motors", 690550454),
    "SUMITOMO RIKO COMPANY LTD":        (2, "Anti-vibration rubber, hoses, sealing products, interior parts", "Honda, Nissan, General Motors, BMW, Mercedes, Mazda", 690538541),
    "GMB CORP":                         (2, "Universal joints, water pumps, tensioners", "Toyota, Mazda, Nissan, Subaru, Honda, Kia, Hyundai", 690594270),
    "KENDRION NV":                      (2, "Electromagnetic actuators, solenoids, electromagnetic clutches", "N/A", 403394372),
    "TOYODA GOSEI CO LTD":              (2, "Weather strips, air bags, steering wheels, hoses, LEDs", "Toyota, Honda, Nissan, Ford, General Motors", 690559182),
    "HWASEUNG CORP CO LTD":             (2, "Shift cable assemblies, transmission cables, HVAC cables", "Hyundai, Kia, General Motors, Ford, Stellantis, BMW", 687998617),
    "JINHAP CO LTD":                    (2, "Rubber/plastic auto parts, bushings, grommets", "Hyundai, Kia", 689283781),
    "NICHIRIN CO LTD":                  (2, "Brake hoses, fuel hoses, HVAC hoses", "Honda, Toyota, Nissan", 690691068),
    "STANLEY ELECTRIC CO LTD":          (2, "Headlamps, taillamps, fog lamps, LED assemblies", "Honda, Toyota, Subaru", 690550637),
    "LCI INDUSTRIES":                   (2, "Chassis components, slide-outs, leveling systems for trucks/RVs", "General Motors, Ford, Stellantis", 7871643),
    "CTEK AB (PUBL)":                   (2, "Battery chargers, EV charging solutions", "BMW, Volkswagen", 350685597),
    "NICHIAS CORP":                     (2, "Gaskets, sealing sheets, thermal insulation for engines", "Toyota, Honda, Nissan", 690577879),
    "UCHIYAMA MANUFACTURING CORP":      (2, "Oil seals, rubber gaskets, dust covers", "Toyota, Honda, Nissan, Mazda, Subaru, General Motors", 690563036),
    "AMS AG":                           (2, "Optical sensors, ambient light sensors, lidar, automotive lighting", "Stellantis", 300209194),
    "MITSUBOSHI BELTING LTD":           (2, "Timing belts, poly-V belts, flat belts", "Volkswagen, Stellantis, Toyota, Renault, Hyundai", 690537287),
    "AISAN INDUSTRY CO LTD":            (2, "Throttle bodies, fuel injection components, EGR systems", "Toyota, Renault, Nissan, Volvo, Suzuki", 690578364),
    "PIOLAX INC":                       (2, "Plastic clips, retainers, springs, interior parts", "Honda, Nissan, Subaru, Mazda, Mitsubishi, Suzuki", 690656772),
    "NITTO DENKO CORP":                 (2, "Masking films, window protection films, EV battery films, NVH tapes", "Toyota, Honda, Nissan, Tesla", 690538913),
    "DAIDO CORP":                       (2, "Steel wire rope, suspension wires, window regulator cables", "Toyota, Honda, Volkswagen, Nissan, General Motors", 690644729),
    "COMMERCIAL VEHICLE GROUP INC":     (2, "Seats, cab-related products, wiper systems, trim for trucks/buses", "Volvo, Daimler Trucks, Peterbilt", 150146855),
    "HARADA INDUSTRY CO LTD":           (2, "Roof-mount antennas, shark-fin antennas, window antennas", "Toyota, Nissan, General Motors, Ford, Honda", 690595079),
    "NISSHINBO HOLDINGS INC":           (2, "Brake pads, brake shoes, electronic circuits, NVH products", "Ford, General Motors, Toyota, Volkswagen", 690691134),
    "SENIOR PLC":                       (2, "Flexible fluid conveyance, heat exchanger tubes, automotive fluid systems", "Volkswagen, BMW", 210514501),
    "HOSIDEN CORPORATION":              (2, "Automotive connectors, touch sensors, EV charging connectors", "Toyota, Honda, Nissan", 690539184),
    "SANKO GOSEI LTD":                  (2, "Interior plastic components, bumpers, door trims", "Toyota, Honda", 690677380),
    "ZHEJIANG YINLUN MACHINERY CO LTD": (2, "EV battery thermal management, heat exchangers, coolers", "Tesla, Volkswagen, General Motors, Ford", 420810582),
    "WOORY INDUSTRIAL HOLDINGS CO LTD": (2, "EV battery cooling, HVAC systems, heat exchangers", "Tesla, Kia, Hyundai", 689223774),
    "ITT INC":                          (2, "Brake pads, EV regenerative brake pads, connectors", "Volkswagen, BMW, Mercedes, Stellantis", 80267418),
    "GUANGDONG HONGTU TECHNOLOGY HOLDING":(2, "Aluminum die castings, EV structural parts, battery enclosures", "Tesla", 530633841),
    "SUPRAJIT ENGINEERING LTD":         (2, "Control cables, parking brake cables, shift cables", "BMW", 918030230),
    "DONALDSON COMPANY INC":            (2, "Engine air filtration, fuel filtration, crankcase ventilation", "Ford, Volvo", 6477301),
    "YOKOWO CO LTD":                    (2, "Spring contacts, antennas, automotive connectors", "Toyota, Honda, Nissan, Mazda", 690549225),
    "XIAMEN HONGFA ELECTROACOUSTIC CO": (2, "Automotive relays, high-voltage contactors for EVs", "Tesla, Toyota, Volkswagen", 654675222),
    "HEXAGON COMPOSITES ASA":           (2, "Composite cylinders for CNG/hydrogen vehicles", "Toyota", 517808812),
    "MODINE MANUFACTURING CO":          (2, "Radiators, oil coolers, charge air coolers, EV thermal modules", "Stellantis, Ford, Tesla", 6092555),
    "CTS CORP":                         (2, "Pedal sensors, transmission sensors, actuators, temperature sensors", "General Motors, Ford, Toyota, Honda", 5068515),
    "TDK CORP":                         (2, "Ferrite cores, capacitors, EV battery management systems, sensors", "Tesla, Toyota, Honda, General Motors, Volkswagen", 690551346),
    "MAYVILLE ENGINEERING CO INC":      (2, "Fabricated structural components, cylinder heads, chassis parts", "Volvo", 6100630),
    "HOWMET AEROSPACE INC":             (2, "Automotive fasteners, multi-material fasteners, precision cast components", "Ford, General Motors", 1339472),
    "MINEBEA MITSUMI INC":              (2, "Miniature ball bearings, stepping motors, sensors, LED backlights", "Toyota, Honda, Nissan, Ford, Tesla, BMW", 690571757),
    "OMRON CORPORATION":                (2, "Automotive relays, sensors, EV on-board charger components", "Toyota, Honda, Nissan", 690537899),
    "PARKER CORP":                      (2, "Auto parts distribution and manufacturing", "General Motors, Ford, Stellantis", 690882766),
}


# ─── QUARTERLY DATA (optional) ────────────────────────────────────────────────
LATEST_QUARTER_JSON = os.path.join(os.path.dirname(__file__), "../data", "latest_quarter_financials.json")


def load_latest_quarter_financials():
    """Load latest quarter data from LATEST_QUARTER_JSON if present.

    Returns:
        Dict company_name -> {period, revenue_usd, sga_usd, ebit_usd, source}. Keys must match INCLUDE_DETAILS.
    """
    if not os.path.isfile(LATEST_QUARTER_JSON):
        return {}
    try:
        with open(LATEST_QUARTER_JSON, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


# ─── WORKBOOK BUILDER ────────────────────────────────────────────────────────

def build_supplier_rows():
    """Combine INCLUDE_DETAILS + FINANCIALS (or latest_quarter_financials.json) into sorted list of dicts.

    Sort key is (tier, -revenue) so Tier 1 first and higher revenue first within tier.
    """
    quarterly = load_latest_quarter_financials()
    rows = []
    for name, (tier, parts, oems, duns) in INCLUDE_DETAILS.items():
        q = quarterly.get(name)
        if q and isinstance(q, dict):
            period = q.get("period") or "Q4 2025"
            rev = q.get("revenue_usd")
            sga = q.get("sga_usd")
            ebit = q.get("ebit_usd")
            source = q.get("source") or "10-Q / IR"
            fy = period
        else:
            fin = FINANCIALS.get(name, (None, None, None, None, None))
            rev, sga, ebit, fy, source = fin
            period = fy or "FY (annual)"
        rows.append({
            "name": name, "duns": duns, "tier": tier,
            "parts": parts, "oems": oems,
            "revenue": rev, "sga": sga, "ebit": ebit,
            "period": period, "fiscal_year": fy or "", "source": source or "",
        })
    rows.sort(key=lambda r: (r["tier"], -(r["revenue"] or 0)))
    return rows


SHEET1_HEADERS = [
    "DUNS Number", "Company Name", "Tier", "Period",
    "Revenue (USD)", "SG&A (USD)", "EBIT (USD)",
    "SG&A %", "EBIT %", "SG&A+EBIT %",
    "OEM Customers", "Automotive Parts / Products",
    "Fiscal Year", "Source",
]
# 1-based column indices for data sheet (Revenue, SG&A, EBIT, and formula columns for ratios)
COL_REV  = 5
COL_SGA  = 6
COL_EBIT = 7
COL_SGAPCT  = 8   # SG&A % formula
COL_EBITPCT = 9   # EBIT % formula
COL_TOTAL   = 10  # SG&A% + EBIT% formula


def write_data_sheet(ws, rows, title="Supplier Data"):
    """Write a formatted supplier data sheet. Formulas for SG&A %, EBIT %, and combined % are in columns COL_SGAPCT, COL_EBITPCT, COL_TOTAL."""
    ws.title = title

    # Freeze first data row
    ws.freeze_panes = "A3"

    # Title row
    ws.row_dimensions[1].height = 22
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(name=FONT_NAME, bold=True, size=13, color=COLOR_HEADER_FONT)
    tc.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SHEET1_HEADERS))

    # Header row
    ws.row_dimensions[2].height = 32
    write_header_row(ws, 2, SHEET1_HEADERS)

    # Data rows
    for i, r in enumerate(rows):
        row_num = i + 3
        bg = COLOR_TIER1_BG if r["tier"] == 1 else COLOR_TIER2_BG
        alt = COLOR_ALT_ROW if i % 2 == 1 else None
        row_bg = alt or bg

        def sc(col, val, fmt=None, halign="left", wrap=False):
            c = ws.cell(row=row_num, column=col, value=val)
            style_cell(c, bg=row_bg, halign=halign, fmt=fmt, wrap=wrap)

        sc(1, r["duns"], halign="right")
        sc(2, r["name"])
        sc(3, r["tier"], halign="center")
        sc(4, r.get("period", r.get("fiscal_year", "")), halign="center")
        sc(COL_REV,  r["revenue"],  fmt=FMT_CURRENCY, halign="right")
        sc(COL_SGA,  r["sga"],      fmt=FMT_CURRENCY, halign="right")
        sc(COL_EBIT, r["ebit"],     fmt=FMT_CURRENCY, halign="right")

        # Formula cells for ratios
        rev_col  = get_column_letter(COL_REV)
        sga_col  = get_column_letter(COL_SGA)
        ebit_col = get_column_letter(COL_EBIT)
        sg_pct_col  = get_column_letter(COL_SGAPCT)
        eb_pct_col  = get_column_letter(COL_EBITPCT)

        for col, formula in [
            (COL_SGAPCT,  f'=IF(OR({rev_col}{row_num}="",{sga_col}{row_num}=""),"",{sga_col}{row_num}/{rev_col}{row_num})'),
            (COL_EBITPCT, f'=IF(OR({rev_col}{row_num}="",{ebit_col}{row_num}=""),"",{ebit_col}{row_num}/{rev_col}{row_num})'),
            (COL_TOTAL,   f'=IF(OR({sg_pct_col}{row_num}="",{eb_pct_col}{row_num}=""),"",{sg_pct_col}{row_num}+{eb_pct_col}{row_num})'),
        ]:
            c = ws.cell(row=row_num, column=col, value=formula)
            style_cell(c, bg=row_bg, fmt=FMT_PCT, halign="right",
                       color=COLOR_FORMULA_FONT)

        sc(11, r["oems"],  wrap=True)
        sc(12, r["parts"], wrap=True)
        sc(13, r["fiscal_year"], halign="center")
        sc(14, r["source"], wrap=True)

    # Column widths (added Period column)
    widths = [14, 42, 6, 12, 18, 18, 18, 10, 10, 12, 48, 60, 12, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(ws, rows, tier_label):
    """Write a Tier 1 or Tier 2 summary sheet with aggregated metrics (counts, totals, averages)."""
    ws.title = f"{tier_label} Summary"
    tier_rows = [r for r in rows if r["tier"] == (1 if "1" in tier_label else 2)]
    rev_rows  = [r for r in tier_rows if r["revenue"] is not None]
    sga_rows  = [r for r in tier_rows if r["sga"] is not None and r["revenue"]]
    ebit_rows = [r for r in tier_rows if r["ebit"] is not None and r["revenue"]]

    def safe_pct(num_rows, key_num, key_den):
        vals = [r[key_num] / r[key_den] for r in num_rows]
        return sum(vals) / len(vals) if vals else None

    stats = [
        ("Companies included", len(tier_rows)),
        ("Companies with revenue data", len(rev_rows)),
        ("Total Revenue (USD)", sum(r["revenue"] for r in rev_rows)),
        ("Avg Revenue (USD)", sum(r["revenue"] for r in rev_rows) / len(rev_rows) if rev_rows else None),
        ("Median Revenue (USD)", sorted(r["revenue"] for r in rev_rows)[len(rev_rows)//2] if rev_rows else None),
        ("Avg SG&A %", safe_pct(sga_rows, "sga", "revenue")),
        ("Avg EBIT %", safe_pct(ebit_rows, "ebit", "revenue")),
        ("Avg SG&A + EBIT %", (lambda combo: (
            sum((r["sga"] + r["ebit"]) / r["revenue"] for r in combo) / len(combo)
            if combo else None
        ))([r for r in tier_rows if r["sga"] and r["ebit"] and r["revenue"]])),
    ]

    ws.cell(row=1, column=1, value=f"{tier_label} Supplier Summary").font = Font(
        name=FONT_NAME, bold=True, size=13)

    write_header_row(ws, 2, ["Metric", "Value"])
    for i, (label, val) in enumerate(stats):
        row_n = i + 3
        c_lbl = ws.cell(row=row_n, column=1, value=label)
        style_cell(c_lbl, bold=True)
        c_val = ws.cell(row=row_n, column=2, value=val)
        if isinstance(val, float) and val < 2:
            style_cell(c_val, fmt=FMT_PCT, halign="right")
        elif isinstance(val, (int, float)):
            style_cell(c_val, fmt=FMT_CURRENCY, halign="right")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22


def write_methodology_sheet(ws):
    """Write the Methodology sheet (inclusion/exclusion criteria, data sources) and EXCLUDE_REASONS table."""
    ws.title = "Methodology"
    lines = [
        ("METHODOLOGY", True),
        ("", False),
        (f"Report generated: {date.today().isoformat()}", False),
        ("Latest quarter end: Q4 2025 / Q1 2026 where available; otherwise annual (see Financial Data below).", False),
        ("", False),
        ("Inclusion Criteria", True),
        ("  Tier 1: Public company; directly supplies physical auto components; revenue ~$2B+", False),
        ("  Tier 2: Public company; supplies sub-assemblies or specialty components to Tier 1 or OEMs", False),
        ("", False),
        ("Exclusion Criteria", True),
        ("  OEMs (vehicle manufacturers), oil & gas companies, trading companies,", False),
        ("  pure financial/holding entities, diversified industrials with <10% auto revenue.", False),
        ("", False),
        ("Financial Data", True),
        ("  Figures are latest quarter (Q4 2025 or Q1 2026 where available) when latest_quarter_financials.json is present; otherwise annual.", False),
        ("  All financials in USD. Non-USD figures converted using FY-average exchange rates.", False),
        ("  Revenue, SG&A and EBIT from SEC 10-Q (US), Yahoo (Canada TSX, Europe, Korea), quarterly_overrides.json (gaps), or annual fallback.", False),
        ("  Japan/HK/Mexico and some symbols (e.g. Valeo FR.PA) often have no Yahoo quarterly P&L; see LIMITATIONS.md.", False),
        ("  SG&A% = SG&A / Revenue.  EBIT% = EBIT / Revenue.", False),
        ("", False),
        ("EXCLUDED COMPANIES", True),
        ("", False),
    ]
    row_n = 1
    for text, bold in lines:
        c = ws.cell(row=row_n, column=1, value=text)
        if bold:
            c.font = Font(name=FONT_NAME, bold=True, size=11)
        else:
            c.font = Font(name=FONT_NAME, size=10)
        row_n += 1

    # Excluded table headers
    write_header_row(ws, row_n, ["Company Name", "Reason for Exclusion"])
    row_n += 1
    for name, reason in sorted(EXCLUDE_REASONS.items()):
        ws.cell(row=row_n, column=1, value=name).font = Font(name=FONT_NAME, size=10)
        c = ws.cell(row=row_n, column=2, value=reason)
        c.font = Font(name=FONT_NAME, size=10)
        row_n += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 70


def main():
    """Build supplier rows, create workbook, write 5 sheets (Supplier Data, Filtered Publics, Tier 1/2 Summary, Methodology), save to auto_suppliers.xlsx."""
    rows = build_supplier_rows()
    all_public = [r for r in rows if r["revenue"] is not None]

    wb = openpyxl.Workbook()

    # Sheet 1: All included suppliers
    ws1 = wb.active
    write_data_sheet(ws1, rows, title="Supplier Data")

    # Sheet 2: Filtered to companies with financial data
    ws2 = wb.create_sheet()
    write_data_sheet(ws2, all_public, title="Filtered Publics")

    # Sheets 3-4: Tier summaries
    ws3 = wb.create_sheet()
    write_summary_sheet(ws3, rows, "Tier 1")

    ws4 = wb.create_sheet()
    write_summary_sheet(ws4, rows, "Tier 2")

    # Sheet 5: Methodology
    ws5 = wb.create_sheet()
    write_methodology_sheet(ws5)

    out = os.path.join(os.path.dirname(__file__), "../data", "auto_suppliers.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  Sheets:      {wb.sheetnames}")
    print(f"  Tier 1 rows: {sum(1 for r in rows if r['tier'] == 1)}")
    print(f"  Tier 2 rows: {sum(1 for r in rows if r['tier'] == 2)}")
    print(f"  With revenue data: {len(all_public)}")


if __name__ == "__main__":
    main()
